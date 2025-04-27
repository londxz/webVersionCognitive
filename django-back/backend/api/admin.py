from django.contrib import admin
from .models import *
import pandas as pd
from django.http import HttpResponse
from django.utils.timezone import localtime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows


@admin.action(description="Export selected results to Excel")
def export_to_excel(modeladmin, request, queryset):
    data = {
        "Имя пользователя": [result.user.username for result in queryset],
        "Возраст": [result.user.age for result in queryset],
        "Образование": [result.user.education for result in queryset],
        "Специальность": [result.user.specialty for result in queryset],
        "Место проживания": [result.user.residence for result in queryset],
        "Рост": [result.user.height for result in queryset],
        "Вес": [result.user.weight for result in queryset],
        "Ведущая рука": [result.user.dominant_hand for result in queryset],
        "Заболевания": [result.user.diseases for result in queryset],
        "Курение": ["Да" if result.user.smoking else "Нет" for result in queryset],
        "Алкоголь": ["Да" if result.user.alcohol else "Нет" for result in queryset],
        "Спорт": ["Да" if result.user.sport else "Нет" for result in queryset],
        "Бессонница": ["Да" if result.user.insomnia else "Нет" for result in queryset],
        "Текущее настроение": [result.user.current_mood for result in queryset],
        "Геймер": ["Да" if result.user.gamer else "Нет" for result in queryset],
        "Название теста": [result.test.test_name for result in queryset],
        "Номер попытки": [result.try_number for result in queryset],
        "Всего ответов": [result.number_all_answers for result in queryset],
        "Правильных ответов": [result.number_correct_answers for result in queryset],
        "Время завершения": [
            localtime(result.complete_time).strftime('%Y-%m-%d %H:%M:%S')
            if result.complete_time else None
            for result in queryset
        ],
        "Точность": [f"{result.accuracy}%" for result in queryset],
    }
    df = pd.DataFrame(data)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=test_results.xlsx'

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=value)

            if r_idx == 1:
                cell.font = Font(bold=True, color="FFFFFF")  # Белый жирный текст
                cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")  # Синий фон
                cell.alignment = Alignment(horizontal="center", vertical="center")  # Выравнивание по центру
                cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                     top=Side(style='thin'), bottom=Side(style='thin'))  # Границы

    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except TypeError:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width

    wb.save(response)
    return response


class TestResultsAdmin(admin.ModelAdmin):
    list_display = ('user', 'test', 'try_number',
                    'number_all_answers',
                    'number_correct_answers', 'accuracy')
    actions = [export_to_excel]


admin.site.register(TestResults, TestResultsAdmin)
admin.site.register(TestNSI)
admin.site.register(User)